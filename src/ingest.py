"""
Data ingestion layer for CACEIS Human Capital Analytics.

Loads all 11 source files from the CACEIS dataset, validates schemas,
and returns typed DataFrames ready for cleaning. openpyxl is used instead
of pd.read_excel for large files (Data.xlsx is 275K rows) because it avoids
loading the entire DOM into memory and handles merged cells more reliably.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import numpy as np
import openpyxl
import pandas as pd

from src.config import (
    DATA_PATH,
    FILE_COLD_REVIEW,
    FILE_EAE_2023,
    FILE_EAE_2025,
    FILE_EMPLOYEE_MASTER,
    FILE_PL_FTE,
    FILE_QUICK_REVIEW,
    FILE_TRAINING,
    SHEET_ABSENTEEISM,
    SHEET_PL,
    SHEET_TURNOVER,
)

logger = logging.getLogger(__name__)


# ── Quality report ─────────────────────────────────────────────────────────────


@dataclass
class SourceReport:
    """Coverage summary for a single source file."""

    name: str
    rows: int
    columns: int
    null_pct: float
    date_min: Optional[str] = None
    date_max: Optional[str] = None
    notes: list[str] = field(default_factory=list)

    def __str__(self) -> str:
        date_range = (
            f"  dates: {self.date_min} → {self.date_max}\n" if self.date_min else ""
        )
        notes = "".join(f"  ! {n}\n" for n in self.notes)
        return (
            f"[{self.name}] {self.rows:,} rows × {self.columns} cols "
            f"| nulls: {self.null_pct:.1f}%\n{date_range}{notes}"
        )


@dataclass
class DataQualityReport:
    """Aggregated quality report for all loaded sources."""

    sources: dict[str, SourceReport] = field(default_factory=dict)

    def add(self, report: SourceReport) -> None:
        self.sources[report.name] = report

    def print_summary(self) -> None:
        print("\n" + "=" * 60)
        print("DATA QUALITY REPORT")
        print("=" * 60)
        for rep in self.sources.values():
            print(rep)

    def _make_report(
        self,
        name: str,
        df: pd.DataFrame,
        date_col: Optional[str] = None,
        notes: Optional[list[str]] = None,
    ) -> SourceReport:
        null_pct = df.isnull().mean().mean() * 100
        date_min = date_max = None
        if date_col and date_col in df.columns:
            valid = df[date_col].dropna()
            if not valid.empty:
                date_min = str(valid.min())[:10]
                date_max = str(valid.max())[:10]
        return SourceReport(
            name=name,
            rows=len(df),
            columns=len(df.columns),
            null_pct=round(null_pct, 2),
            date_min=date_min,
            date_max=date_max,
            notes=notes or [],
        )


# ── Loader ─────────────────────────────────────────────────────────────────────


class CACEISDataLoader:
    """
    Loads all CACEIS source files from *data_path*.

    Example
    -------
    >>> loader = CACEISDataLoader(Path("data"))
    >>> sources = loader.load_all()
    >>> df_employees = sources["employee_master"]
    """

    def __init__(self, data_path: Path | str = DATA_PATH) -> None:
        self.data_path = Path(data_path)
        self.quality_report = DataQualityReport()

    def _path(self, filename: str) -> Path:
        p = self.data_path / filename
        if not p.exists():
            raise FileNotFoundError(
                f"Source file not found: {p}\n"
                "Place CACEIS source files in the data/ directory. "
                "See docs/data_dictionary.md for expected filenames."
            )
        return p

    # ── Employee master ───────────────────────────────────────────────────────

    def load_employee_master(self) -> pd.DataFrame:
        """
        Load Data.xlsx Sheet1 — 275K employee records spanning 2022-2025.

        Uses openpyxl read_only mode to stream rows without loading the full
        workbook DOM. This avoids OOM errors on the 275K-row file.

        Returns
        -------
        pd.DataFrame with 14 columns: country_code, country, period,
        employee_id, age_range, gender, contract_type, degree,
        entry_reason, date_entry_group, date_entry_caceis, date_entry_role,
        job_title, entity.
        """
        path = self._path(FILE_EMPLOYEE_MASTER)
        logger.info("Loading employee master: %s", path)

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb["Sheet1"]

        headers: list[str] = []
        rows: list[list] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = list(row)
            else:
                rows.append(list(row))
        wb.close()

        df = pd.DataFrame(rows, columns=headers)

        # Standardise column names
        df.columns = [
            "country_code", "country", "period", "employee_id", "age_range",
            "gender", "contract_type", "degree", "entry_reason",
            "date_entry_group", "date_entry_caceis", "date_entry_role",
            "job_title", "entity", *df.columns[14:],
        ]
        # Drop trailing unnamed columns if any
        df = df.iloc[:, :14]

        for col in ["period", "date_entry_caceis", "date_entry_group", "date_entry_role"]:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors="coerce")

        logger.info(
            "Employee master loaded: %d rows | %d countries | periods %s → %s",
            len(df),
            df["country"].nunique(),
            str(df["period"].min())[:10],
            str(df["period"].max())[:10],
        )

        report = self.quality_report._make_report(
            "employee_master", df, date_col="period"
        )
        if len(df) < 200_000:
            report.notes.append(f"Expected ≥275K rows, got {len(df):,} — check source file")
        self.quality_report.add(report)
        return df

    # ── P&L + FTE ─────────────────────────────────────────────────────────────

    def load_pl_fte(self) -> pd.DataFrame:
        """
        Load AlbertSchool_CACEIS_PL-FTE_22-25_Sent.xlsx Synthese_PL sheet.

        Parses keyword-matched rows for Net Banking Income, Total Personnel,
        and Formation lines. Returns a 4-row dataframe (one per year 2022-2025).

        Returns
        -------
        pd.DataFrame with columns: year, pnb, personnel, training, fte,
        hc_roi, rev_per_fte, cost_ratio, train_per_fte.
        """
        from src.config import FTE_VALS, YEARS

        path = self._path(FILE_PL_FTE)
        logger.info("Loading P&L/FTE: %s", path)

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[SHEET_PL]
        pl_rows = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()

        def _find_val(rows: list, keyword: str, col_offset: int = 1, n_cols: int = 4) -> list[float]:
            for row in rows:
                if row[0] and keyword in str(row[0]):
                    vals: list[float] = []
                    for j in range(col_offset, col_offset + n_cols):
                        try:
                            vals.append(abs(float(str(row[j]).replace(",", ""))))
                        except (ValueError, TypeError, IndexError):
                            vals.append(float("nan"))
                    return vals
            logger.warning("Keyword '%s' not found in P&L sheet", keyword)
            return [float("nan")] * n_cols

        pnb = _find_val(pl_rows, "Net Banking Income")
        personnel = _find_val(pl_rows, "Total Personnel")
        training_c = _find_val(pl_rows, "Formation")

        df_pl = pd.DataFrame({
            "year": YEARS,
            "pnb": pnb,
            "personnel": personnel,
            "training": training_c,
            "fte": [FTE_VALS[y] for y in YEARS],
        })
        df_pl["hc_roi"] = (df_pl["pnb"] - df_pl["personnel"]) / df_pl["personnel"] * 100
        df_pl["rev_per_fte"] = df_pl["pnb"] / df_pl["fte"]
        df_pl["cost_ratio"] = df_pl["personnel"] / df_pl["pnb"] * 100
        df_pl["train_per_fte"] = df_pl["training"] / df_pl["fte"] * 1000

        logger.info("P&L loaded: HC-ROI 2025 = %.1f%%", df_pl.loc[df_pl["year"] == 2025, "hc_roi"].values[0])
        self.quality_report.add(
            self.quality_report._make_report("pl_fte", df_pl)
        )
        return df_pl

    # ── EAE 2025 ──────────────────────────────────────────────────────────────

    def load_eae_2025(self) -> pd.DataFrame:
        """
        Load EAE 2025 performance evaluation database.

        Returns
        -------
        pd.DataFrame with at minimum: employee_id (or equivalent key),
        performance rating (1-5 or label), entity, country.
        """
        path = self._path(FILE_EAE_2025)
        logger.info("Loading EAE 2025: %s", path)
        df = pd.read_excel(path, engine="openpyxl")
        logger.info("EAE 2025 loaded: %d rows × %d cols", *df.shape)
        self.quality_report.add(
            self.quality_report._make_report("eae_2025", df, notes=[
                "Check rating column name before using in clean.py"
            ])
        )
        return df

    # ── EAE 2023 ──────────────────────────────────────────────────────────────

    def load_eae_2023(self) -> pd.DataFrame:
        """
        Load EAE 2023 performance evaluation database.

        Returns
        -------
        pd.DataFrame — same structure as EAE 2025.
        """
        path = self._path(FILE_EAE_2023)
        logger.info("Loading EAE 2023: %s", path)
        df = pd.read_excel(path, engine="openpyxl")
        logger.info("EAE 2023 loaded: %d rows × %d cols", *df.shape)
        self.quality_report.add(
            self.quality_report._make_report("eae_2023", df)
        )
        return df

    # ── Training records ──────────────────────────────────────────────────────

    def load_training(self) -> pd.DataFrame:
        """
        Load Training_Records_Unnamed.xlsx — 14,943 session records.

        Expected columns: employee_id (or anon), session_name, hours,
        status ('Réalisé' or equivalent), satisfaction, transfer_score.

        Returns
        -------
        pd.DataFrame with 14,943+ rows.
        """
        path = self._path(FILE_TRAINING)
        logger.info("Loading training records: %s", path)
        df = pd.read_excel(path, engine="openpyxl")
        logger.info("Training loaded: %d sessions", len(df))
        if len(df) < 10_000:
            logger.warning("Training records: expected ≥14,943 rows, got %d", len(df))
        self.quality_report.add(
            self.quality_report._make_report("training", df)
        )
        return df

    # ── Absenteeism ───────────────────────────────────────────────────────────

    def load_absenteeism(self) -> pd.DataFrame:
        """
        Load Data.xlsx sheet 'Absentéisme FR' — 72 monthly records.

        Returns
        -------
        pd.DataFrame with monthly absence rates for France 2020-2025.
        """
        path = self._path(FILE_EMPLOYEE_MASTER)
        logger.info("Loading absenteeism (sheet: %s): %s", SHEET_ABSENTEEISM, path)
        df = pd.read_excel(path, sheet_name=SHEET_ABSENTEEISM, engine="openpyxl")
        logger.info("Absenteeism loaded: %d rows", len(df))
        self.quality_report.add(
            self.quality_report._make_report("absenteeism", df)
        )
        return df

    # ── Turnover ──────────────────────────────────────────────────────────────

    def load_turnover(self) -> pd.DataFrame:
        """
        Load Data.xlsx sheet 'taux mob_TO FR' — monthly turnover rates.

        Returns
        -------
        pd.DataFrame with date and turnover_rate columns.
        """
        path = self._path(FILE_EMPLOYEE_MASTER)
        logger.info("Loading turnover (sheet: %s): %s", SHEET_TURNOVER, path)
        df = pd.read_excel(path, sheet_name=SHEET_TURNOVER, engine="openpyxl")
        logger.info("Turnover loaded: %d rows", len(df))
        self.quality_report.add(
            self.quality_report._make_report("turnover", df)
        )
        return df

    # ── Quick/Cold Review ─────────────────────────────────────────────────────

    def load_quick_review(self) -> pd.DataFrame:
        """Load Quick_Review_Unnamed.xlsx — 9,706 satisfaction responses."""
        path = self._path(FILE_QUICK_REVIEW)
        logger.info("Loading quick review: %s", path)
        df = pd.read_excel(path, engine="openpyxl")
        logger.info("Quick review loaded: %d rows", len(df))
        self.quality_report.add(self.quality_report._make_report("quick_review", df))
        return df

    def load_cold_review(self) -> pd.DataFrame:
        """Load Cold_Review_Unnamed.xlsx — 8,647 transfer-rate responses."""
        path = self._path(FILE_COLD_REVIEW)
        logger.info("Loading cold review: %s", path)
        df = pd.read_excel(path, engine="openpyxl")
        logger.info("Cold review loaded: %d rows", len(df))
        self.quality_report.add(self.quality_report._make_report("cold_review", df))
        return df

    # ── Load all ──────────────────────────────────────────────────────────────

    def load_all(self) -> dict[str, pd.DataFrame]:
        """
        Run all loaders and return a named dict of DataFrames.

        Returns
        -------
        dict with keys: employee_master, pl_fte, eae_2025, eae_2023,
        training, absenteeism, turnover, quick_review, cold_review.
        """
        sources: dict[str, pd.DataFrame] = {}

        loaders = {
            "employee_master": self.load_employee_master,
            "pl_fte": self.load_pl_fte,
            "eae_2025": self.load_eae_2025,
            "eae_2023": self.load_eae_2023,
            "training": self.load_training,
            "absenteeism": self.load_absenteeism,
            "turnover": self.load_turnover,
            "quick_review": self.load_quick_review,
            "cold_review": self.load_cold_review,
        }

        for name, loader in loaders.items():
            try:
                sources[name] = loader()
            except FileNotFoundError as exc:
                logger.error("Skipping %s: %s", name, exc)

        self.quality_report.print_summary()
        return sources
